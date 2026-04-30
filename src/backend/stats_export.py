import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

CLR_TITLE_BG = "D6E4F0"
CLR_TITLE_FONT = "1A3A52"
CLR_STUDENT_BG = "F0F4F8"
CLR_STUDENT_FONT = "1A3A52"
CLR_BLOCK_BG = "E8F0FE"
CLR_BLOCK_FONT = "1A3A52"
CLR_TOTAL_BG = "D6EAD6"
CLR_TOTAL_FONT = "1E4620"
CLR_ROW_BG = "FFFFFF"
CLR_DONE_FULL = "C6EFCE"  # зелёный — 100%
CLR_DONE_HIGH = "FFEB9C"  # жёлтый — 50–99%
CLR_DONE_LOW = "FFC7CE"  # красный — < 50%
CLR_BORDER = "C5D5E8"


def _thin_border():
    s = Side(style="thin", color=CLR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _cell_color(done, total):
    if total == 0:
        return CLR_ROW_BG
    pct = done / total
    if pct >= 1.0:
        return CLR_DONE_FULL
    if pct >= 0.5:
        return CLR_DONE_HIGH
    return CLR_DONE_LOW


def build_course_stats_xlsx(db, course_id: int) -> tuple[bytes, str]:
    """
    Строит xlsx-отчёт и возвращает (bytes, filename).
    db  — соединение sqlite3 с row_factory = sqlite3.Row
    """
    #Данные курса
    course = db.execute("SELECT * FROM courses WHERE id=?", (course_id,)).fetchone()
    if not course:
        raise ValueError(f"Курс {course_id} не найден")

    #модули
    blocks_raw = db.execute(
        "SELECT * FROM course_blocks WHERE course_id=? ORDER BY order_index",
        (course_id,)
    ).fetchall()

    # Для каждого блока — список items (lesson_id или task_id)
    blocks = []
    for b in blocks_raw:
        items = db.execute(
            "SELECT * FROM block_items WHERE block_id=? ORDER BY order_index",
            (b["id"],)
        ).fetchall()
        blocks.append({
            "id": b["id"],
            "title": b["title"],
            "items": [dict(it) for it in items],
        })

    # Если блоков нет — один псевдоблок из всех уроков (legacy)
    if not blocks:
        lessons = db.execute(
            "SELECT id, title FROM lessons WHERE course_id=? ORDER BY order_index",
            (course_id,)
        ).fetchall()
        pseudo_items = [
            {"type": "lesson", "lesson_id": l["id"], "task_id": None, "title": l["title"]}
            for l in lessons
        ]
        blocks = [{"id": 0, "title": "Уроки", "items": pseudo_items}]

    #Студенты
    students = db.execute(
        """SELECT u.id, u.login, u.first_name, u.last_name
           FROM users u
           JOIN user_courses uc ON u.id = uc.user_id
           WHERE uc.course_id = ?
           ORDER BY u.last_name, u.first_name""",
        (course_id,)
    ).fetchall()

    #одним запросом выгружаем все завершённые уроки / задания
    done_lessons = set()  # (user_id, lesson_id)
    for row in db.execute(
            """SELECT up.user_id, up.lesson_id
               FROM user_progress up
               JOIN lessons l ON up.lesson_id = l.id
               WHERE l.course_id = ? AND up.completed = 1""",
            (course_id,)
    ):
        done_lessons.add((row["user_id"], row["lesson_id"]))

    done_tasks = set()  # (user_id, task_id)
    for row in db.execute(
            """SELECT ta.user_id, ta.task_id
               FROM task_answers ta
               JOIN tasks t ON ta.task_id = t.id
               WHERE t.course_id = ? AND ta.is_correct = 1""",
            (course_id,)
    ):
        done_tasks.add((row["user_id"], row["task_id"]))

    # 5. Вспомогательная функция: сколько элементов блока выполнил студент
    def block_progress(user_id, block_items):
        done = total = 0
        for it in block_items:
            if it["type"] == "lesson" and it.get("lesson_id"):
                total += 1
                if (user_id, it["lesson_id"]) in done_lessons:
                    done += 1
            elif it["type"] == "task" and it.get("task_id"):
                total += 1
                if (user_id, it["task_id"]) in done_tasks:
                    done += 1
        return done, total

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"

    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center")

    # заголовок файла
    total_cols = 1 + len(blocks) + 1
    last_col_letter = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = f"Статистика прохождения курса: {course['title']}"
    title_cell.font = Font(name="Arial", bold=True, size=13, color=CLR_TITLE_FONT)
    title_cell.fill = _fill(CLR_TITLE_BG)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # шапка
    header_row = 2
    c = ws.cell(row=header_row, column=1, value="Студент")
    c.font = Font(name="Arial", bold=True, size=10, color=CLR_STUDENT_FONT)
    c.fill = _fill(CLR_STUDENT_BG)
    c.alignment = center
    c.border = border

    for col_idx, block in enumerate(blocks, start=2):
        c = ws.cell(row=header_row, column=col_idx, value=block["title"])
        c.font = Font(name="Arial", bold=True, size=10, color=CLR_BLOCK_FONT)
        c.fill = _fill(CLR_BLOCK_BG)
        c.alignment = center
        c.border = border

    total_col = total_cols
    c = ws.cell(row=header_row, column=total_col, value="Итого")
    c.font = Font(name="Arial", bold=True, size=10, color=CLR_TOTAL_FONT)
    c.fill = _fill(CLR_TOTAL_BG)
    c.alignment = center
    c.border = border
    ws.row_dimensions[header_row].height = 36

    # студенты
    for s_idx, student in enumerate(students):
        row = header_row + 1 + s_idx
        bg = CLR_ROW_BG

        # Столбец A — имя
        name = f"{student['last_name']} {student['first_name']}".strip() or student["login"]
        nc = ws.cell(row=row, column=1, value=name)
        nc.font = Font(name="Arial", size=10)
        nc.fill = _fill(bg)
        nc.alignment = left_al
        nc.border = border

        total_done = total_all = 0

        # По одному столбцу на блок
        for col_idx, block in enumerate(blocks, start=2):
            done, total = block_progress(student["id"], block["items"])
            total_done += done
            total_all += total

            if total == 0:
                val = "—"
                cell_bg = bg
            else:
                pct = int(done / total * 100)
                val = f"{done} / {total} ({pct}%)"
                cell_bg = _cell_color(done, total)

            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = _fill(cell_bg)
            c.alignment = center_wrap
            c.border = border

        # Итоговый столбец
        if total_all == 0:
            total_val = "—"
            total_bg = bg
        else:
            pct = int(total_done / total_all * 100)
            total_val = f"{total_done} / {total_all} ({pct}%)"
            total_bg = _cell_color(total_done, total_all)

        tc = ws.cell(row=row, column=total_col, value=total_val)
        tc.font = Font(name="Arial", size=10, bold=True)
        tc.fill = _fill(total_bg)
        tc.alignment = center_wrap
        tc.border = border

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len * 1.2 + 4, 80)

    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row[0].row].height = max(20, max_lines * 15)
    ws.freeze_panes = "B3"

    # Сериализация в байты
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = "".join(c for c in course["title"] if c.isalnum() or c in " _-").strip()
    filename = f"{safe_title}_{date.today().strftime('%Y-%m-%d')}.xlsx"

    return buf.read(), filename
