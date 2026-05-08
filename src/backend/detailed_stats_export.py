import io
import json
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


CLR_TITLE_BG     = "D6E4F0"
CLR_TITLE_FONT   = "1A3A52"
CLR_STUDENT_BG   = "F0F4F8"
CLR_STUDENT_FONT = "1A3A52"
CLR_BLOCK_BG     = "E8F0FE"
CLR_BLOCK_FONT   = "1A3A52"
CLR_TASK_BG      = "F5F8FF"
CLR_TASK_FONT    = "1A3A52"
CLR_TOTAL_BG     = "D6EAD6"
CLR_TOTAL_FONT   = "1E4620"
CLR_CORRECT      = "CCFFCC"
CLR_WRONG        = "FFCCCC"
CLR_PENDING      = "FFF3CD"
CLR_NO_ANSWER    = "EAEAEA"
CLR_BORDER       = "C5D5E8"


FIXED_WIDTH_TYPES = {"text", "match"}
FIXED_COL_WIDTH   = 28
FIXED_ROW_HEIGHT  = 55
DEFAULT_ROW_HEIGHT = 20


def _thin_border():
    s = Side(style="thin", color=CLR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _format_answer(answer_text: str, task_type: str) -> str:
    """Форматирует ответ студента для отображения в ячейке."""
    if not answer_text:
        return ""
    if task_type == "match":
        try:
            items = json.loads(answer_text)
            return "\n".join(f"{i+1}. {v}" for i, v in enumerate(items))
        except Exception:
            return answer_text
    return answer_text


def build_detailed_stats_xlsx(db, course_id: int) -> tuple[bytes, str]:
    """
    Строит подробный xlsx-отчёт и возвращает (bytes, filename).
    db — соединение sqlite3 с row_factory = sqlite3.Row
    """
    #Данные курса
    course = db.execute("SELECT * FROM courses WHERE id=?", (course_id,)).fetchone()
    if not course:
        raise ValueError(f"Курс {course_id} не найден")

    #Блоки с заданиями
    blocks_raw = db.execute(
        "SELECT * FROM course_blocks WHERE course_id=? ORDER BY order_index",
        (course_id,)
    ).fetchall()

    blocks = []
    for b in blocks_raw:
        items = db.execute(
            """SELECT bi.*, t.question, t.task_type
               FROM block_items bi
               JOIN tasks t ON bi.task_id = t.id
               WHERE bi.block_id = ? AND bi.type = 'task'
               ORDER BY bi.order_index""",
            (b["id"],)
        ).fetchall()
        task_items = [dict(it) for it in items]
        if task_items:
            blocks.append({"id": b["id"], "title": b["title"], "tasks": task_items})

    #нет блоков — берём все задания курса
    if not blocks:
        tasks_raw = db.execute(
            "SELECT id, question, task_type FROM tasks WHERE course_id=? ORDER BY id",
            (course_id,)
        ).fetchall()
        blocks = [{"id": 0, "title": "Задания", "tasks": [
            {"task_id": t["id"], "question": t["question"], "task_type": t["task_type"]}
            for t in tasks_raw
        ]}]

    all_tasks = []
    for block in blocks:
        for task in block["tasks"]:
            all_tasks.append({**task, "block_title": block["title"]})

    if not all_tasks:
        raise ValueError("В курсе нет заданий")

    #студенты курса
    students = db.execute(
        """SELECT u.id, u.login, u.first_name, u.last_name
           FROM users u
           JOIN user_courses uc ON u.id = uc.user_id
           WHERE uc.course_id = ?
           ORDER BY u.last_name, u.first_name""",
        (course_id,)
    ).fetchall()

    #Все ответы одним запросом
    answers = {}
    for row in db.execute(
        """SELECT ta.user_id, ta.task_id, ta.answer_text, ta.is_correct
           FROM task_answers ta
           JOIN tasks t ON ta.task_id = t.id
           WHERE t.course_id = ?""",
        (course_id,)
    ):
        answers[(row["user_id"], row["task_id"])] = {
            "text":       row["answer_text"] or "",
            "is_correct": row["is_correct"],
        }


    wb = Workbook()
    ws = wb.active
    ws.title = "Подробный отчёт"

    border      = _thin_border()
    center      = Alignment(horizontal="center", vertical="center", wrap_text=False)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    left_al     = Alignment(horizontal="left",   vertical="center")

    total_cols  = 1 + len(all_tasks) + 1
    last_letter = get_column_letter(total_cols)

    #заголовок
    ws.merge_cells(f"A1:{last_letter}1")
    c = ws["A1"]
    c.value     = f"Подробный отчёт по курсу: {course['title']}"
    c.font      = Font(name="Arial", bold=True, size=13, color=CLR_TITLE_FONT)
    c.fill      = _fill(CLR_TITLE_BG)
    c.alignment = center
    ws.row_dimensions[1].height = 30

    #названия блоков
    col_cursor = 2
    for block in blocks:
        span         = len(block["tasks"])
        start_letter = get_column_letter(col_cursor)
        end_letter   = get_column_letter(col_cursor + span - 1)
        if span > 1:
            ws.merge_cells(f"{start_letter}2:{end_letter}2")
        c = ws.cell(row=2, column=col_cursor, value=block["title"])
        c.font      = Font(name="Arial", bold=True, size=10, color=CLR_BLOCK_FONT)
        c.fill      = _fill(CLR_BLOCK_BG)
        c.alignment = center
        c.border    = border
        col_cursor += span

    ws.cell(row=2, column=1).fill   = _fill(CLR_STUDENT_BG)
    ws.cell(row=2, column=1).border = border
    ws.cell(row=2, column=total_cols).fill   = _fill(CLR_TOTAL_BG)
    ws.cell(row=2, column=total_cols).border = border
    ws.row_dimensions[2].height = 36

    # названия заданий
    c = ws.cell(row=3, column=1, value="Студент")
    c.font      = Font(name="Arial", bold=True, size=10, color=CLR_STUDENT_FONT)
    c.fill      = _fill(CLR_STUDENT_BG)
    c.alignment = center
    c.border    = border

    for col_idx, task in enumerate(all_tasks, start=2):
        c = ws.cell(row=3, column=col_idx, value=task["question"])
        c.font      = Font(name="Arial", size=9, color=CLR_TASK_FONT)
        c.fill      = _fill(CLR_TASK_BG)
        c.alignment = center_wrap
        c.border    = border

    c = ws.cell(row=3, column=total_cols, value="Итого")
    c.font      = Font(name="Arial", bold=True, size=10, color=CLR_TOTAL_FONT)
    c.fill      = _fill(CLR_TOTAL_BG)
    c.alignment = center
    c.border    = border
    ws.row_dimensions[3].height = 48

    # студенты
    for s_idx, student in enumerate(students):
        row  = 4 + s_idx
        name = f"{student['last_name']} {student['first_name']}".strip() or student["login"]

        nc = ws.cell(row=row, column=1, value=name)
        nc.font      = Font(name="Arial", size=10)
        nc.fill      = _fill(CLR_STUDENT_BG)
        nc.alignment = left_al
        nc.border    = border

        correct_count    = 0
        has_fixed_height = False

        for col_idx, task in enumerate(all_tasks, start=2):
            task_id   = task["task_id"]
            task_type = task["task_type"]
            answer    = answers.get((student["id"], task_id))

            if answer is None:
                bg  = CLR_NO_ANSWER
                val = ""
                al  = center_wrap
            elif answer["is_correct"] == 1:
                correct_count += 1
                bg  = CLR_CORRECT
                val = _format_answer(answer["text"], task_type)
                al  = left_wrap if task_type in FIXED_WIDTH_TYPES else center_wrap
            elif answer["is_correct"] == 0:
                bg  = CLR_WRONG
                val = _format_answer(answer["text"], task_type)
                al  = left_wrap if task_type in FIXED_WIDTH_TYPES else center_wrap
            else:
                # NULL — текстовое задание ещё не проверено преподавателем
                bg  = CLR_PENDING
                val = "На проверке"
                al  = center_wrap

            if task_type in FIXED_WIDTH_TYPES:
                has_fixed_height = True

            c = ws.cell(row=row, column=col_idx, value=val)
            c.font      = Font(name="Arial", size=9)
            c.fill      = _fill(bg)
            c.alignment = al
            c.border    = border

        # Итоговый столбец
        total_tasks = len(all_tasks)
        pct = int(correct_count / total_tasks * 100) if total_tasks else 0
        tc = ws.cell(row=row, column=total_cols,
                     value=f"{correct_count} / {total_tasks} ({pct}%)")
        tc.font      = Font(name="Arial", size=10, bold=True)
        tc.fill      = _fill(CLR_TOTAL_BG)
        tc.alignment = center_wrap
        tc.border    = border

        ws.row_dimensions[row].height = (
            FIXED_ROW_HEIGHT if has_fixed_height else DEFAULT_ROW_HEIGHT
        )

    max_name = max(
        (len(f"{s['last_name']} {s['first_name']}".strip() or s["login"])
         for s in students),
        default=10
    )
    ws.column_dimensions["A"].width = min(max_name * 1.2 + 4, 40)

    for col_idx, task in enumerate(all_tasks, start=2):
        col_letter = get_column_letter(col_idx)
        if task["task_type"] in FIXED_WIDTH_TYPES:
            ws.column_dimensions[col_letter].width = FIXED_COL_WIDTH
        else:
            ws.column_dimensions[col_letter].width = min(
                len(task["question"]) * 1.2 + 4, 40
            )

    ws.column_dimensions[get_column_letter(total_cols)].width = 18


    ws.freeze_panes = "B4"


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = "".join(c for c in course["title"] if c.isalnum() or c in " _-").strip()
    filename   = f"Подробный_отчёт_{safe_title}_{date.today().strftime('%Y-%m-%d')}.xlsx"

    return buf.read(), filename
